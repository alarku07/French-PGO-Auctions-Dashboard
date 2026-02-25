"""EEX data parser service.

Handles:
- ZIP archive extraction + Excel file parsing (backfill)
- HTML scraping for latest auction results (daily sync)
- HTML scraping for auction calendar (daily sync)
"""
import asyncio
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import httpx
import openpyxl
import structlog
from bs4 import BeautifulSoup, Tag
from openpyxl.worksheet.worksheet import Worksheet

logger = structlog.get_logger("parser")


def _normalize_col_name(s: str) -> str:
    """Strip non-ASCII characters and lowercase for encoding-agnostic matching.

    EEX Excel files sometimes encode accented characters (é, €) as multi-byte
    sequences that don't round-trip cleanly through openpyxl.  Removing all
    non-ASCII characters before comparison makes matching robust regardless of
    the encoding variant actually stored in the file.
    """
    return re.sub(r"[^\x00-\x7f]", "", s.strip().lower())


# Known column name variants → canonical field name
# Includes bilingual French/English headers as used by EEX Excel files.
_COLUMN_MAP: dict[str, str] = {
    # Region (English and bilingual French/English)
    "region": "region",
    "regions": "region",
    "région / region": "region",
    "r\ufffdgion / region": "region",  # encoding fallback for older files
    # Technology (bilingual)
    "technology": "technology",
    "energy source": "technology",
    "technologie / technology": "technology",
    # Volume offered / auctioned
    "volume offered": "volume_offered_mwh",
    "volume offered (mwh)": "volume_offered_mwh",
    "volume offered mwh": "volume_offered_mwh",
    "offered volume": "volume_offered_mwh",
    "offered volume (mwh)": "volume_offered_mwh",
    "total volume auctionned": "volume_offered_mwh",  # EEX spelling in files
    "total volume auctioned": "volume_offered_mwh",
    # Volume allocated / sold
    "volume allocated": "volume_allocated_mwh",
    "volume allocated (mwh)": "volume_allocated_mwh",
    "volume awarded": "volume_allocated_mwh",
    "volume awarded (mwh)": "volume_allocated_mwh",
    "awarded volume": "volume_allocated_mwh",
    "total volume sold": "volume_allocated_mwh",
    # Price
    "weighted average price": "weighted_avg_price_eur",
    "weighted average price (eur/mwh)": "weighted_avg_price_eur",
    "weighted avg price": "weighted_avg_price_eur",
    "average clearing price": "weighted_avg_price_eur",
    "clearing price": "weighted_avg_price_eur",
    "weighted average price (\u20ac / mwh)": "weighted_avg_price_eur",  # € symbol
    "weighted average price (? / mwh)": "weighted_avg_price_eur",  # encoding fallback
    # Bids / winners
    "number of bids": "num_bids",
    "no. of bids": "num_bids",
    "bids": "num_bids",
    "number of winning bids": "num_winning_bids",
    "winning bids": "num_winning_bids",
    "number of winners per couple region/technology": "num_winning_bids",
    "number of winners per couple region/tech": "num_winning_bids",
    # Production period
    "production period": "production_period",
    "production month": "production_period",
    # Auction date
    "auction date": "auction_date",
    "date": "auction_date",
    # Reserve / floor price
    "reserve price": "reserve_price_eur",
    "floor price": "reserve_price_eur",
    "reserve price (eur/mwh)": "reserve_price_eur",
}

# Normalized version of _COLUMN_MAP (non-ASCII stripped) for robust lookups.
# Built once at module load time so it's not recomputed per-cell.
_COLUMN_MAP_NORM: dict[str, str] = {
    _normalize_col_name(k): v for k, v in _COLUMN_MAP.items()
}

# French technology name → canonical English name
_TECHNOLOGY_ALIASES: dict[str, str] = {
    # English variants
    "wind": "Wind",
    "onshore wind": "Wind",
    "offshore wind": "Wind",
    "hydro": "Hydro",
    "hydropower": "Hydro",
    "solar": "Solar",
    "pv": "Solar",
    "photovoltaic": "Solar",
    "thermal": "Thermal",
    "biomass": "Thermal",
    # French variants as used in EEX files
    "eolien onshore": "Wind",
    "\u00e9olien onshore": "Wind",
    "eolien offshore": "Wind",
    "\u00e9olien offshore": "Wind",
    "hydraulique": "Hydro",
    "solaire": "Solar",
    "thermique": "Thermal",
}

# Canonical French region names.  We normalise inbound strings by stripping
# non-ASCII characters and matching against these canonical values (also
# stripped).  This fixes encoding corruption in some EEX file vintages where
# accented chars appear as U+FFFD replacement characters.
_CANONICAL_REGIONS: list[str] = [
    "Auvergne-Rh\u00f4ne-Alpes",  # Rhône
    "Bourgogne-Franche-Comt\u00e9",  # Comté
    "Bretagne",
    "Centre-Val de Loire",
    "Grand Est",
    "Haut-de-France",
    "Normandie",
    "Nouvelle-Aquitaine",
    "Occitanie",
    "Pays de la Loire",
    "Provence-Alpes-C\u00f4te d\u2019Azur",  # Côte / apostrophe
    "Provence-Alpes-C\u00f4te d'Azur",  # Côte / straight quote
    "\u00cele-de-France",  # Île
]

# Lookup: ASCII-stripped lowercase → canonical name
_REGION_NORM_MAP: dict[str, str] = {
    _normalize_col_name(r): r for r in _CANONICAL_REGIONS
}


def _normalise_region(raw: str | None) -> str | None:
    """Return the canonical region name, fixing encoding artefacts if needed."""
    if raw is None:
        return None
    text = raw.strip()
    if not text:
        return None
    canonical = _REGION_NORM_MAP.get(_normalize_col_name(text))
    return canonical if canonical is not None else text


def _normalise_technology(raw: str | None) -> str | None:
    """Normalise technology string to canonical form (English)."""
    if raw is None:
        return None
    cleaned = raw.strip().lower()
    # Exact match first (handles bilingual French names)
    if cleaned in _TECHNOLOGY_ALIASES:
        return _TECHNOLOGY_ALIASES[cleaned]
    # Substring match for English names
    for alias, canonical in _TECHNOLOGY_ALIASES.items():
        if alias in cleaned:
            return canonical
    return None  # unknown technology — will be stored as NULL


def _compute_aggregate_rows(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Compute technology=NULL aggregate rows from per-technology records.

    Groups by (auction_date, region, production_period) and sums volumes.
    Weighted-average price is computed as SUM(price * vol) / SUM(vol).
    These aggregate rows are required by the stats and chart queries.
    """
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}
    for r in records:
        key = (r.get("auction_date"), r.get("region"), r.get("production_period"))
        if None in key:
            continue
        if key not in groups:
            groups[key] = {
                "volume_offered_mwh": Decimal(0),
                "volume_allocated_mwh": Decimal(0),
                "price_x_vol": Decimal(0),
                "vol_with_price": Decimal(0),
                "status": r.get("status", "past"),
            }
        g = groups[key]
        if r.get("volume_offered_mwh") is not None:
            g["volume_offered_mwh"] += r["volume_offered_mwh"]
        if r.get("volume_allocated_mwh") is not None:
            g["volume_allocated_mwh"] += r["volume_allocated_mwh"]
            if r.get("weighted_avg_price_eur") is not None:
                g["price_x_vol"] += (
                    r["weighted_avg_price_eur"] * r["volume_allocated_mwh"]
                )
                g["vol_with_price"] += r["volume_allocated_mwh"]

    agg: list[dict[str, Any]] = []
    for (auction_date, region, production_period), g in groups.items():
        wavg = (
            g["price_x_vol"] / g["vol_with_price"] if g["vol_with_price"] else None
        )
        agg.append({
            "auction_date": auction_date,
            "region": region,
            "production_period": production_period,
            "technology": None,
            "status": g["status"],
            "volume_offered_mwh": g["volume_offered_mwh"] or None,
            "volume_allocated_mwh": g["volume_allocated_mwh"] or None,
            "weighted_avg_price_eur": wavg,
        })
    return agg


def _parse_decimal(value: Any) -> Decimal | None:
    """Safely parse a cell value to Decimal, return None on failure."""
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except InvalidOperation:
        return None


def _parse_int(value: Any) -> int | None:
    """Safely parse a cell value to int, treating '-' and blanks as None."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    try:
        return int(float(text))
    except (ValueError, OverflowError):
        return None


def _parse_date_value(value: Any) -> date | None:
    """Parse a date from various Excel cell formats."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_production_period(value: Any) -> str | None:
    """Parse production period to 'YYYY-MM' format."""
    if value is None:
        return None
    text = str(value).strip()
    # Already YYYY-MM
    if re.match(r"^\d{4}-\d{2}$", text):
        return text
    # Try full date
    parsed = _parse_date_value(text)
    if parsed:
        return parsed.strftime("%Y-%m")
    # "January 2025" or "Jan 2025"
    for fmt in ("%B %Y", "%b %Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return text  # return as-is and let the caller handle it


# ─── ZIP / Excel parsing ──────────────────────────────────────────────────────


def _detect_header_row(ws: Worksheet) -> int | None:
    """Scan for the header row by looking for known column names."""
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=20, values_only=True), 1
    ):
        for cell in row:
            if cell is not None and _normalize_col_name(str(cell)) in _COLUMN_MAP_NORM:
                return row_idx
    return None


def _parse_worksheet(
    ws: Worksheet,
    filename: str,
    default_auction_date: date | None = None,
    default_production_period: str | None = None,
) -> list[dict[str, Any]]:
    """Parse a single worksheet into a list of auction record dicts."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        logger.warning(
            "unrecognised_column_layout",
            filename=filename,
            sheet=ws.title,
        )
        return []

    # Build column index → canonical field mapping
    header_cells = list(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )[0]
    col_map: dict[int, str] = {}
    for col_idx, cell in enumerate(header_cells):
        if cell is not None:
            canonical = _COLUMN_MAP_NORM.get(_normalize_col_name(str(cell)))
            if canonical:
                col_map[col_idx] = canonical

    if not col_map:
        logger.warning(
            "no_recognised_columns",
            filename=filename,
            sheet=ws.title,
        )
        return []

    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None for cell in row):
            continue  # skip empty rows

        record: dict[str, Any] = {}
        for col_idx, field in col_map.items():
            if col_idx < len(row):
                record[field] = row[col_idx]

        # Apply defaults for fields not in this file
        if "auction_date" not in record or record["auction_date"] is None:
            record["auction_date"] = default_auction_date
        if "production_period" not in record or record["production_period"] is None:
            record["production_period"] = default_production_period

        # Type coercions
        record["auction_date"] = _parse_date_value(record.get("auction_date"))
        record["production_period"] = _parse_production_period(
            record.get("production_period")
        )
        record["technology"] = _normalise_technology(record.get("technology"))
        record["volume_offered_mwh"] = _parse_decimal(record.get("volume_offered_mwh"))
        record["volume_allocated_mwh"] = _parse_decimal(
            record.get("volume_allocated_mwh")
        )
        record["weighted_avg_price_eur"] = _parse_decimal(
            record.get("weighted_avg_price_eur")
        )
        record["reserve_price_eur"] = _parse_decimal(record.get("reserve_price_eur"))
        record["num_bids"] = _parse_int(record.get("num_bids"))
        record["num_winning_bids"] = _parse_int(record.get("num_winning_bids"))
        record["region"] = _normalise_region(record.get("region"))

        # Skip rows with no usable key fields
        if record.get("auction_date") is None or not record.get("region"):
            continue

        record["status"] = "past"
        records.append(record)

    return records + _compute_aggregate_rows(records)


def _parse_excel_bytes(
    data: bytes,
    filename: str,
    default_auction_date: date | None = None,
    default_production_period: str | None = None,
) -> list[dict[str, Any]]:
    """Parse an .xlsx file from bytes."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("excel_open_failed", filename=filename, error=str(exc))
        return []

    # Extract date hints from filename: YYYYMMDD_MonthName_Year_N_...
    if default_auction_date is None or default_production_period is None:
        name = Path(filename).stem
        date_match = re.match(r"^(\d{8})_(\w+)_(\d{4})", name)
        if date_match:
            try:
                default_auction_date = default_auction_date or datetime.strptime(
                    date_match.group(1), "%Y%m%d"
                ).date()
                month_str = f"{date_match.group(2)} {date_match.group(3)}"
                default_production_period = (
                    default_production_period
                    or datetime.strptime(month_str, "%B %Y").strftime("%Y-%m")
                )
            except ValueError:
                pass

    records: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_records = _parse_worksheet(
            ws, filename, default_auction_date, default_production_period
        )
        records.extend(sheet_records)

    return records


async def parse_zip_archive(zip_path: Path) -> list[dict[str, Any]]:
    """Extract a ZIP archive and parse all .xlsx files within it.

    Runs openpyxl in a thread pool to avoid blocking the event loop.
    """
    def _sync_parse() -> list[dict[str, Any]]:
        all_records: list[dict[str, Any]] = []
        with ZipFile(zip_path, "r") as zf:
            xlsx_files = [
                name for name in zf.namelist() if name.lower().endswith(".xlsx")
            ]
            logger.info("zip_contents", path=str(zip_path), xlsx_count=len(xlsx_files))

            for name in xlsx_files:
                try:
                    data = zf.read(name)
                    records = _parse_excel_bytes(data, name)
                    all_records.extend(records)
                    logger.debug("parsed_xlsx", filename=name, records=len(records))
                except Exception as exc:
                    logger.warning("xlsx_parse_error", filename=name, error=str(exc))

        return all_records

    return await asyncio.to_thread(_sync_parse)


# ─── HTML scraping — latest results ──────────────────────────────────────────


def _extract_table_rows(
    table: BeautifulSoup,
) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) from a BeautifulSoup table element."""
    headers: list[str] = []
    rows: list[list[str]] = []

    thead = table.find("thead")
    if isinstance(thead, Tag):
        header_row = thead.find("tr")
        if isinstance(header_row, Tag):
            headers = [
                th.get_text(strip=True).lower()
                for th in header_row.find_all(["th", "td"])
            ]

    tbody_found = table.find("tbody")
    tbody: Tag = tbody_found if isinstance(tbody_found, Tag) else table
    for tr in tbody.find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
        if cells and any(c for c in cells):
            rows.append(cells)

    return headers, rows


async def scrape_latest_results(
    eex_base_url: str,
    http_client: httpx.AsyncClient,
) -> list[dict[str, Any]]:
    """Scrape the latest auction results from the EEX French Auctions page.

    Returns a list of auction record dicts with status='past'.
    """
    url = f"{eex_base_url}/en/markets/energy-certificates/french-auctions-power"
    try:
        response = await http_client.get(url, follow_redirects=True, timeout=30.0)
        response.raise_for_status()
    except httpx.HTTPError as exc:
        logger.error("eex_fetch_failed", url=url, error=str(exc))
        raise

    soup = BeautifulSoup(response.text, "html.parser")

    records: list[dict[str, Any]] = []

    # Find tables that look like result tables (containing region or volume columns)
    tables = soup.find_all("table")
    if not tables:
        logger.warning("no_tables_found", url=url)
        return records

    # Try to extract auction_date from page context (heading or caption)
    auction_date = _extract_auction_date_from_page(soup)

    for table in tables:
        headers, rows = _extract_table_rows(table)
        if not headers:
            continue

        # Detect result tables by presence of known columns
        col_indices: dict[str, int] = {}
        for i, h in enumerate(headers):
            canonical = _COLUMN_MAP.get(h)
            if canonical:
                col_indices[canonical] = i

        if "region" not in col_indices and "technology" not in col_indices:
            continue  # not an auction result table

        for row in rows:
            record: dict[str, Any] = {"auction_date": auction_date, "status": "past"}
            for field, idx in col_indices.items():
                if idx < len(row):
                    record[field] = row[idx] or None

            # Coerce types
            record["auction_date"] = auction_date
            record["production_period"] = _parse_production_period(
                record.get("production_period")
            )
            record["region"] = _normalise_region(record.get("region"))
            record["technology"] = _normalise_technology(record.get("technology"))
            record["volume_offered_mwh"] = _parse_decimal(
                record.get("volume_offered_mwh")
            )
            record["volume_allocated_mwh"] = _parse_decimal(
                record.get("volume_allocated_mwh")
            )
            record["weighted_avg_price_eur"] = _parse_decimal(
                record.get("weighted_avg_price_eur")
            )

            if record.get("region") and record.get("auction_date"):
                records.append(record)

    logger.info("scraped_latest_results", count=len(records), url=url)
    return records


def _extract_auction_date_from_page(soup: BeautifulSoup) -> date | None:
    """Attempt to extract the most recent auction date from page headings."""
    # Look for date patterns in headings or paragraphs near result tables
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "caption", "p"]):
        text = tag.get_text(strip=True)
        # Pattern: "18 November 2025" or "2025-11-18"
        for pattern, fmt in [
            (r"\d{1,2}\s+\w+\s+\d{4}", "%d %B %Y"),
            (r"\d{4}-\d{2}-\d{2}", "%Y-%m-%d"),
        ]:
            match = re.search(pattern, text)
            if match:
                try:
                    return datetime.strptime(match.group(0), fmt).date()
                except ValueError:
                    continue
    return None


# ─── HTML scraping — auction calendar ────────────────────────────────────────


async def scrape_auction_calendar(
    eex_base_url: str,
    http_client: httpx.AsyncClient,
) -> list[dict[str, Any]]:
    """Scrape the upcoming auction calendar from the EEX French Auctions page.

    Returns a list of auction record dicts with status='upcoming'.
    """
    url = f"{eex_base_url}/en/markets/energy-certificates/french-auctions-power"
    try:
        response = await http_client.get(url, follow_redirects=True, timeout=30.0)
        response.raise_for_status()
    except httpx.HTTPError as exc:
        logger.error("calendar_fetch_failed", url=url, error=str(exc))
        raise

    soup = BeautifulSoup(response.text, "html.parser")
    records: list[dict[str, Any]] = []

    tables = soup.find_all("table")
    for table in tables:
        headers, rows = _extract_table_rows(table)
        if not headers:
            continue

        # Calendar tables contain order book columns
        header_text = " ".join(headers)
        if "order" not in header_text and "matching" not in header_text:
            continue

        # Map column indices
        col_map: dict[str, int] = {}
        calendar_cols = {
            "auctioning month": "auction_date",
            "auction month": "auction_date",
            "auction date": "auction_date",
            "production month": "production_period",
            "production period": "production_period",
            "order book opening": "order_book_open",
            "order book open": "order_book_open",
            "order book closure": "order_book_close",
            "order book close": "order_book_close",
            "order matching": "order_matching",
            "order book matching": "order_matching",
        }
        for i, h in enumerate(headers):
            for keyword, field in calendar_cols.items():
                if keyword in h:
                    col_map[field] = i
                    break

        if "auction_date" not in col_map:
            continue

        for row in rows:
            record: dict[str, Any] = {"status": "upcoming"}
            for field, idx in col_map.items():
                if idx < len(row) and row[idx]:
                    record[field] = row[idx]

            parsed_date = _parse_date_value(record.get("auction_date"))
            if parsed_date is None:
                continue

            # Skip dates in the past
            if parsed_date < date.today():
                continue

            record["auction_date"] = parsed_date
            record["production_period"] = _parse_production_period(
                record.get("production_period")
            )
            record["order_book_open"] = _parse_datetime_loose(
                record.get("order_book_open")
            )
            record["order_book_close"] = _parse_datetime_loose(
                record.get("order_book_close")
            )
            record["order_matching"] = _parse_datetime_loose(
                record.get("order_matching")
            )
            record.setdefault("region", "France")  # calendar may lack per-region rows
            record["technology"] = None

            if record.get("production_period"):
                records.append(record)

    logger.info("scraped_calendar", count=len(records), url=url)
    return records


def _parse_datetime_loose(value: Any) -> datetime | None:
    """Parse a loose date/datetime string from the EEX calendar table."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Try ISO datetime first
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d %B %Y %H:%M",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None
